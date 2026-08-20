from __future__ import annotations

import hashlib
import json
import math
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCES = {
    "rijing": {
        "name": "日井",
        "current_parent": ROOT / "work/r_parent_aug.xlsx",
        "prior_parent": ROOT / "work/r_parent_jul.xlsx",
        "current_shop": ROOT / "work/r_shop_aug.xlsx",
        "prior_shop": ROOT / "work/r_shop_jul.xlsx",
        "files": [
            ("16gu6yR92YA02oICGS86A_1hjja_DLIyQ", "parentskudetail.20260701_20260731.xlsx", "2026/07/01—07/31", "2026-08-20T08:33:12Z", "prior_parent"),
            ("1_mjj5IlZ6aGbLSvLk4ZQuuYnlos494IT", "vincenthung.shopee-shop-stats.20260701-20260731.xlsx", "2026/07/01—07/31", "2026-08-20T08:30:34Z", "prior_shop"),
            ("1mz5TRv7f8BqHm-bBoVmciMud8_ErbNbB", "parentskudetail.20260801_20260819.xlsx", "2026/08/01—08/19", "2026-08-20T08:34:12Z", "current_parent"),
            ("1kUakLkLBHKkHRi4HjSiIHTawl-MBZAjt", "vincenthung.shopee-shop-stats.20260801-20260819.xlsx", "2026/08/01—08/19", "2026-08-20T08:31:48Z", "current_shop"),
        ],
    },
    "wenxin": {
        "name": "文信",
        "current_parent": ROOT / "work/w_parent_aug.xlsx",
        "prior_parent": ROOT / "work/w_parent_jul.xlsx",
        "current_shop": ROOT / "work/w_shop_aug.xlsx",
        "prior_shop": ROOT / "work/w_shop_jul.xlsx",
        "files": [
            ("18P8zmEL_YeZRHsLw-fNZHxxn4D5R1I9l", "parentskudetail.20260701_20260731 (1).xlsx", "2026/07/01—07/31", "2026-08-20T08:33:47Z", "prior_parent"),
            ("1PCcj5UyMgksLm8p2gI_ryZogPB1sujcy", "fumioh.shopee-shop-stats.20260701-20260731.xlsx", "2026/07/01—07/31", "2026-08-20T08:31:13Z", "prior_shop"),
            ("17NSOoBuBZPibQMLsjTZfyJREY2UzsAwm", "parentskudetail.20260801_20260819 (1).xlsx", "2026/08/01—08/19", "2026-08-20T08:35:03Z", "current_parent"),
            ("18UUrWio_DbUcyLh6xMdsJi4U4by5Ry7f", "fumioh.shopee-shop-stats.20260801-20260819.xlsx", "2026/08/01—08/19", "2026-08-20T08:32:16Z", "current_shop"),
        ],
    },
}

PRODUCT_COLUMNS = {
    "revenue": "銷售額(全部訂單) (TWD)",
    "impressions": "商品曝光次數",
    "clicks": "商品點擊數",
    "orders": "全部訂單",
    "visitors": "商品訪客數",
    "page_views": "商品頁瀏覽數",
    "bounces": "跳出商品的不重複訪客",
    "cart_visitors": "商品頁訪客數(加入購物車)",
    "cart_items": "加入購物車(件數)",
    "buyers": "買家(全部訂單)",
}

SHOP_COLUMNS = {
    "revenue": "總銷售額 (TWD)",
    "net_revenue": "銷售額 （扣除平台補貼）",
    "orders": "訂單總數",
    "aov": "平均訂單金額",
    "clicks": "商品點擊數",
    "visitors": "訪客數",
    "order_conversion": "訂單轉換率",
    "invalid_orders": "不成立的訂單",
    "invalid_revenue": "不成立訂單的銷售額",
    "refund_orders": "退貨/退款訂單",
    "refund_revenue": "退貨/退款的銷售額",
    "buyers": "買家數",
    "new_buyers": "新買家數",
    "returning_buyers": "舊買家數",
    "potential_buyers": "潛在買家數",
    "repurchase_rate": "回購率",
}


def number(value):
    if value in (None, "", "-"):
        return 0.0
    text = str(value).replace(",", "").strip()
    try:
        return float(text.rstrip("%")) / (100 if text.endswith("%") else 1)
    except ValueError:
        return 0.0


def ratio(a, b):
    return a / b if b else 0


def compact_name(name, limit=46):
    name = re.sub(r"^[🌸★☆*\s]+", "", name)
    name = re.sub(r"(鋐宇泵浦|鋐宇)", "", name)
    name = re.sub(r"\s+", " ", name).strip("★☆* -")
    return name if len(name) <= limit else name[: limit - 1] + "…"


def percentile(values, current):
    return sum(value <= current for value in values) / len(values) if values else 0


def read_products(path):
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True)["最佳表現商品"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    index = {name: headers.index(name) for name in headers}
    products = []
    for row in rows:
        if str(row[index["商品規格ID"]]) != "-":
            continue
        item = {
            "id": str(row[index["商品ID"]]),
            "name": compact_name(str(row[index["商品名稱"]])),
            "fullName": str(row[index["商品名稱"]]),
        }
        item.update({key: number(row[index[source]]) for key, source in PRODUCT_COLUMNS.items()})
        item.update(
            {
                "ctr": ratio(item["clicks"], item["impressions"]),
                "conversion": ratio(item["orders"], item["visitors"]),
                "cart_rate": ratio(item["cart_visitors"], item["visitors"]),
                "bounce_rate": ratio(item["bounces"], item["visitors"]),
            }
        )
        products.append(item)
    return products


def read_shop(path):
    ws = openpyxl.load_workbook(path, read_only=True, data_only=True)["全部訂單"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    index = {name: headers.index(name) for name in headers}
    summary = next(row for row in rows[1:] if re.fullmatch(r"\d{4}-\d{2}-\d{2}-\d{4}-\d{2}-\d{2}", str(row[0])))
    daily_rows = [row for row in rows[1:] if re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(row[0]))]

    def parse(row):
        parsed = {key: number(row[index[source]]) for key, source in SHOP_COLUMNS.items()}
        parsed["date"] = str(row[0])
        return parsed

    totals = parse(summary)
    daily = [parse(row) for row in daily_rows]
    totals["visitor_to_potential"] = ratio(totals["potential_buyers"], totals["visitors"])
    totals["potential_to_buyer"] = ratio(totals["buyers"], totals["potential_buyers"])
    totals["visitor_to_buyer"] = ratio(totals["buyers"], totals["visitors"])
    totals["orders_per_buyer"] = ratio(totals["orders"], totals["buyers"])
    totals["valid_order_rate"] = 1 - ratio(totals["invalid_orders"], totals["orders"])
    totals["refund_rate"] = ratio(totals["refund_orders"], totals["orders"])
    return {"totals": {key: round(value, 4) if isinstance(value, float) else value for key, value in totals.items()}, "daily": daily}


def score_products(products):
    pools = {
        "traffic": [math.log1p(product["visitors"]) for product in products],
        "revenue": [math.log1p(product["revenue"]) for product in products],
        "ctr": [product["ctr"] for product in products],
        "conversion": [product["conversion"] for product in products],
        "cart": [product["cart_rate"] for product in products],
    }
    weights = {"revenue": 27, "conversion": 24, "traffic": 18, "ctr": 16, "cart": 15}
    for product in products:
        components = {
            "revenue": round(percentile(pools["revenue"], math.log1p(product["revenue"])) * 100),
            "conversion": round(percentile(pools["conversion"], product["conversion"]) * 100),
            "traffic": round(percentile(pools["traffic"], math.log1p(product["visitors"])) * 100),
            "ctr": round(percentile(pools["ctr"], product["ctr"]) * 100),
            "cart": round(percentile(pools["cart"], product["cart_rate"]) * 100),
        }
        product["components"] = components
        product["score"] = round(sum(components[key] * weights[key] for key in weights) / 100)


def aggregate_products(products):
    totals = {metric: sum(product[metric] for product in products) for metric in PRODUCT_COLUMNS}
    totals.update(
        {
            "ctr": ratio(totals["clicks"], totals["impressions"]),
            "conversion": ratio(totals["orders"], totals["visitors"]),
            "cart_rate": ratio(totals["cart_visitors"], totals["visitors"]),
            "bounce_rate": ratio(totals["bounces"], totals["visitors"]),
        }
    )
    return {key: round(value, 4) for key, value in totals.items()}


def volume_comparison(current, prior, current_days, prior_days):
    keys = ["revenue", "orders", "clicks", "visitors", "buyers", "potential_buyers"]
    result = {}
    for key in keys:
        current_daily = current[key] / current_days
        prior_daily = prior[key] / prior_days
        result[key] = round((current_daily - prior_daily) / prior_daily, 4) if prior_daily else None
    return result


def tactics(shop, products):
    no_order = sorted(
        [product for product in products if product["orders"] == 0 and product["visitors"] >= 10],
        key=lambda product: (product["visitors"], product["cart_visitors"]),
        reverse=True,
    )
    target = no_order[0]["name"] if no_order else "高流量未成交商品"
    if shop == "rijing":
        return [
            {"title": "先修有訪客但零訂單商品", "body": f"「{target}」已有瀏覽卻未成交。首圖補上適用場景、規格判斷與到貨承諾，並把安裝疑問前移。", "effort": "60–90 分鐘"},
            {"title": "合併同質商品的決策資訊", "body": "建立型號比較表，依用途、揚程、功率、保固與現貨狀態引導選型，降低消費者來回搜尋。", "effort": "1–2 小時"},
            {"title": "高單價商品補強信任", "body": "補上實機影片、尺寸／噪音／保固、出貨方式與售後流程，降低下單前的不確定感。", "effort": "2 小時"},
        ]
    return [
        {"title": "複製冠軍商品資訊結構", "body": "把冠軍商品清楚的尺寸標示、MIT／保固與快速出貨訊號，複製到同系列其他商品首屏。", "effort": "1–2 小時"},
        {"title": "處理有加購卻零成交", "body": f"「{target}」已有明確流量。檢查規格命名、相容型號、運費與到貨日是否造成結帳疑慮。", "effort": "45–60 分鐘"},
        {"title": "降低冠軍商品集中風險", "body": "用商品頁內的替代型號比較與使用情境連結，把冠軍商品流量導向第二、三名商品。", "effort": "60 分鐘"},
    ]


def checkpoint(source, item):
    file_id, file_name, period, modified, path_key = item
    path = source[path_key]
    return {
        "fileId": file_id,
        "fileName": file_name,
        "period": period,
        "modified": modified,
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
    }


def build_shop(key, source):
    products = read_products(source["current_parent"])
    prior_products = read_products(source["prior_parent"])
    current_shop = read_shop(source["current_shop"])
    prior_shop = read_shop(source["prior_shop"])
    score_products(products)
    product_totals = aggregate_products(products)
    prior_product_totals = aggregate_products(prior_products)
    totals = current_shop["totals"]
    prior_totals = prior_shop["totals"]
    by_score = sorted(products, key=lambda product: (product["score"], product["revenue"]), reverse=True)
    by_revenue = sorted(products, key=lambda product: product["revenue"], reverse=True)
    by_traffic = sorted(products, key=lambda product: (product["visitors"], product["clicks"]), reverse=True)
    by_opportunity = sorted(
        [product for product in products if product["orders"] == 0],
        key=lambda product: (product["visitors"], product["cart_visitors"]),
        reverse=True,
    )
    files = [checkpoint(source, item) for item in source["files"]]
    return {
        "name": source["name"],
        "date": "2026/08/01—08/19",
        "priorDate": "2026/07/01—07/31",
        "source": {"modified": max(item["modified"] for item in files), "files": files},
        "coverage": {"parentProducts": len(products), "priorParentProducts": len(prior_products), "periodDays": 19, "priorPeriodDays": 31},
        "totals": totals,
        "priorTotals": prior_totals,
        "productTotals": product_totals,
        "priorProductTotals": prior_product_totals,
        "comparisonDaily": volume_comparison(totals, prior_totals, 19, 31),
        "history": [
            {"period": "2026/07", "days": 31, "totals": prior_totals},
            {"period": "08/01—08/19", "days": 19, "totals": totals},
        ],
        "daily": current_shop["daily"],
        "topProducts": by_score[:6],
        "revenueLeaders": by_revenue[:7],
        "trafficLeaders": by_traffic[:7],
        "opportunities": by_opportunity[:6],
        "actions": tactics(key, products),
        "table": by_revenue[:30],
    }


payload = {
    "version": "1.2.0",
    "generatedAt": "2026-08-20T17:00:00+08:00",
    "method": "Four-file incremental checkpoints per shop; shop totals from Shopee shop stats; product analysis from parent SKU summary rows; volume comparison normalized to daily averages.",
    "competitiveness": {
        "scope": "同一店舖、同一期間的父商品內部相對比較；不是外部市場或蝦皮官方評分。",
        "weights": {"revenue": 27, "conversion": 24, "traffic": 18, "ctr": 16, "cart": 15},
    },
    "shops": {key: build_shop(key, source) for key, source in SOURCES.items()},
}

output = ROOT / "data/analysis.json"
output.parent.mkdir(parents=True, exist_ok=True)
output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(output)
